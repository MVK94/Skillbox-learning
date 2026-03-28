from openpyxl import Workbook
wb = Workbook()
wb.create_sheet(title = "Первый лист", index = 0)
sheet = wb["Первый лист"]

sheet.row_dimensions[1].height = 150
sheet.column_dimensions['B'].width = 100

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Объединенные ячейки'

wb.save("test.xlsx")